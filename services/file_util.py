import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def create_excel(headers, rows):
    wb = Workbook()
    ws = wb.active
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws['{}1'.format(col_letter)] = str(header)

    for row_num, row in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row, 1):
            col_letter = get_column_letter(col_num)
            ws['{}'.format(col_letter) + str(row_num)] = str(cell_value)

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file
